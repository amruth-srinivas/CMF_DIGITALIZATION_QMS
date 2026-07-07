from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
from urllib.parse import urlparse
from datetime import time
import os
import io
import csv
from pydantic import BaseModel
import pdfplumber
from docx import Document as DocxDocument
import openpyxl


from DB.database import get_db
from DB.minio_client import get_minio_client
from DB.models.oms import (
    Operation as OperationModel,
    OperationDocument as OperationDocumentModel,
    ToolWithPart as ToolWithPartModel,
    PartType as PartTypeModel
)
from DB.models.configuration import WorkCenter as WorkCenterModel, Machine as MachineModel
from DB.schemas.oms import Operation, OperationCreate, OperationUpdate

router = APIRouter(
    prefix="/operations",
    tags=["operations"]
)


class OperationPreview(BaseModel):
    operation_number: str
    operation_name: str
    setup_time: Optional[str] = None
    cycle_time: Optional[str] = None
    work_instructions: Optional[str] = None
    notes: Optional[str] = None


def _normalize_header(name: str) -> str:
    return name.strip().lower().replace("\n", " ").replace("\r", " ")


def _match_column(header: str) -> Optional[str]:
    h = _normalize_header(header)
    if "op" in h and ("number" in h or "num" in h or "#" in h):
        return "operation_number"
    if "operation" in h and "name" in h:
        return "operation_name"
    if "setup" in h:
        return "setup_time"
    if "cycle" in h:
        return "cycle_time"
    if "instruction" in h or "work inst" in h:
        return "work_instructions"
    if "note" in h:
        return "notes"
    return None


def _find_header_row(rows: list) -> int:
    """
    Scan rows top-to-bottom and return the index of the first row
    that contains BOTH an op-number column and an operation-name column.
    Returns -1 if not found.
    This lets us skip title rows, meta-info rows, blank rows etc.
    """
    for idx, row in enumerate(rows):
        mapped = set()
        for cell in row:
            key = _match_column(str(cell) if cell is not None else "")
            if key:
                mapped.add(key)
        if "operation_number" in mapped and "operation_name" in mapped:
            return idx
    return -1


def _parse_rows(rows: list) -> List[OperationPreview]:
    if not rows:
        return []

    # Find the actual header row (skip title / meta rows at top)
    header_idx = _find_header_row(rows)
    if header_idx == -1:
        return []

    header = rows[header_idx]
    header_map: dict[str, int] = {}
    for col_idx, cell in enumerate(header):
        key = _match_column(str(cell) if cell is not None else "")
        if key:
            header_map[key] = col_idx

    if "operation_number" not in header_map or "operation_name" not in header_map:
        return []

    result: List[OperationPreview] = []
    for row in rows[header_idx + 1:]:
        # Pad row to at least the length of the header
        cells = [(str(c) if c is not None else "").strip() for c in row]
        while len(cells) <= max(header_map.values()):
            cells.append("")

        # Skip completely empty rows and meta/note rows that slip through
        if not "".join(cells).strip():
            continue

        op_num  = cells[header_map["operation_number"]]
        op_name = cells[header_map["operation_name"]]

        # Skip rows where op_number or op_name look like note/footer lines
        if not op_num or not op_name:
            continue

        data = {
            "operation_number":  op_num,
            "operation_name":    op_name,
            "setup_time":        cells[header_map["setup_time"]]        if "setup_time"        in header_map else None,
            "cycle_time":        cells[header_map["cycle_time"]]        if "cycle_time"        in header_map else None,
            "work_instructions": cells[header_map["work_instructions"]] if "work_instructions" in header_map else None,
            "notes":             cells[header_map["notes"]]             if "notes"             in header_map else None,
        }
        # Replace empty strings with None for optional fields
        for k in ("setup_time", "cycle_time", "work_instructions", "notes"):
            if data[k] == "":
                data[k] = None

        result.append(OperationPreview(**data))

    return result


# ── CSV ───────────────────────────────────────────────────────────────────────
def _parse_csv(content: bytes) -> List[OperationPreview]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.reader(io.StringIO(text))
    # Keep every row (including blanks) so _find_header_row can locate the header
    rows = list(reader)
    return _parse_rows(rows)


# ── XLSX ──────────────────────────────────────────────────────────────────────
def _parse_xlsx(content: bytes) -> List[OperationPreview]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        parsed = _parse_rows(rows)
        if parsed:
            return parsed
    return []


# ── DOCX ──────────────────────────────────────────────────────────────────────
def _parse_docx(content: bytes) -> List[OperationPreview]:
    doc = DocxDocument(io.BytesIO(content))
    for table in doc.tables:
        rows = [[cell.text for cell in row.cells] for row in table.rows]
        parsed = _parse_rows(rows)
        if parsed:
            return parsed
    return []


# ── PDF ───────────────────────────────────────────────────────────────────────
def _parse_pdf(content: bytes) -> List[OperationPreview]:
    operations: List[OperationPreview] = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                parsed = _parse_rows(table)
                if parsed:
                    operations.extend(parsed)
            if operations:
                break
    return operations


@router.post("/", response_model=Operation, status_code=status.HTTP_201_CREATED)
def create_operation(operation: OperationCreate, db: Session = Depends(get_db)):
    """Create a new operation"""
    data = operation.model_dump()
    part_type_id = data.get("part_type_id") or 1
    data["part_type_id"] = part_type_id

    # Validate required times only for non Out-Source operations
    setup_time_val = data.get("setup_time")
    cycle_time_val = data.get("cycle_time")
    zero_time = time(0, 0, 0)
    if part_type_id != 2:
        if (
            not setup_time_val
            or not cycle_time_val
            or setup_time_val == zero_time
            or cycle_time_val == zero_time
        ):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="setup_time and cycle_time are mandatory and cannot be 00:00:00 for non Out-Source operations",
            )

    # Ensure part_id is present
    part_id = data.get("part_id")
    if not part_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="part_id is required for operations",
        )

    # Handle operation_number: ensure uniqueness per part and auto-generate when missing
    op_number_raw = data.get("operation_number")
    op_number = op_number_raw.strip() if isinstance(op_number_raw, str) else None

    if op_number:
        existing = (
            db.query(OperationModel)
            .filter(
                OperationModel.part_id == part_id,
                OperationModel.operation_number == op_number,
            )
            .first()
        )
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Operation number '{op_number}' already exists for this part",
            )
        data["operation_number"] = op_number
    else:
        existing_ops = (
            db.query(OperationModel)
            .filter(OperationModel.part_id == part_id)
            .all()
        )
        max_num = 0
        for op in existing_ops:
            try:
                n = int(str(op.operation_number).strip())
            except (TypeError, ValueError):
                continue
            if n > max_num:
                max_num = n
        next_num = max_num + 10 if max_num > 0 else 10
        data["operation_number"] = str(next_num)

    if part_type_id == 2:
        if not data.get("from_date") or not data.get("to_date"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Outsource operations require from_date and to_date",
            )
    db_operation = OperationModel(**data)
    db.add(db_operation)
    db.commit()
    db.refresh(db_operation)
    db_operation = (
        db.query(OperationModel)
        .options(joinedload(OperationModel.user))
        .filter(OperationModel.id == db_operation.id)
        .first()
    )
    pt = db.query(PartTypeModel).filter(PartTypeModel.id == db_operation.part_type_id).first()
    db_operation.part_type_name = pt.type_name if pt else None
    return db_operation


@router.post("/bulk", response_model=List[Operation], status_code=status.HTTP_201_CREATED)
def create_operations_bulk(operations: List[OperationCreate], db: Session = Depends(get_db)):
    """Create many operations in one request (same validations as single create)."""
    if not operations:
        return []

    part_ids = {op.part_id for op in operations if op.part_id}
    if not part_ids or len(part_ids) != 1:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="bulk create requires exactly one part_id across all operations",
        )
    part_id = next(iter(part_ids))

    existing_ops = db.query(OperationModel).filter(OperationModel.part_id == part_id).all()
    max_num = 0
    for op in existing_ops:
        try:
            n = int(str(op.operation_number).strip())
        except (TypeError, ValueError):
            continue
        max_num = max(max_num, n)
    next_num = max_num + 10 if max_num > 0 else 10

    requested_numbers: set[str] = set()
    created_ids: List[int] = []

    try:
        for op_in in operations:
            data = op_in.model_dump()
            pt_id = data.get("part_type_id") or 1
            data["part_type_id"] = pt_id

            setup_time_val = data.get("setup_time")
            cycle_time_val = data.get("cycle_time")
            zero_time = time(0, 0, 0)
            if pt_id != 2:
                if (
                    not setup_time_val
                    or not cycle_time_val
                    or setup_time_val == zero_time
                    or cycle_time_val == zero_time
                ):
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="setup_time and cycle_time are mandatory and cannot be 00:00:00 for non Out-Source operations",
                    )

            if not data.get("part_id"):
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="part_id is required for operations")

            if pt_id == 2:
                if not data.get("from_date") or not data.get("to_date"):
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Outsource operations require from_date and to_date",
                    )

            op_number_raw = data.get("operation_number")
            op_number = op_number_raw.strip() if isinstance(op_number_raw, str) else None
            if op_number:
                if op_number in requested_numbers:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Duplicate operation number '{op_number}' in bulk request",
                    )
                existing = (
                    db.query(OperationModel)
                    .filter(OperationModel.part_id == part_id, OperationModel.operation_number == op_number)
                    .first()
                )
                if existing:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Operation number '{op_number}' already exists for this part",
                    )
                data["operation_number"] = op_number
                requested_numbers.add(op_number)
            else:
                while str(next_num) in requested_numbers:
                    next_num += 10
                data["operation_number"] = str(next_num)
                requested_numbers.add(str(next_num))
                next_num += 10

            db_operation = OperationModel(**data)
            db.add(db_operation)
            db.flush()
            created_ids.append(db_operation.id)

        db.commit()

        created = (
            db.query(OperationModel)
            .options(joinedload(OperationModel.user))
            .filter(OperationModel.id.in_(created_ids))
            .order_by(OperationModel.id.asc())
            .all()
        )
        pt_ids = {o.part_type_id for o in created if o.part_type_id is not None}
        pt_map = {}
        if pt_ids:
            pts = db.query(PartTypeModel).filter(PartTypeModel.id.in_(pt_ids)).all()
            pt_map = {p.id: p.type_name for p in pts}
        for op in created:
            op.part_type_name = pt_map.get(op.part_type_id)
        return created

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create operations: {str(e)}")


@router.get("/", response_model=List[Operation])
def get_operations(user_id: int | None = None, db: Session = Depends(get_db)):
    query = (
        db.query(OperationModel)
        .options(joinedload(OperationModel.user))
        .order_by(OperationModel.id.asc())
    )
    if user_id is not None:
        query = query.filter(OperationModel.user_id == user_id)
    operations = query.all()

    work_center_ids = {op.workcenter_id for op in operations if op.workcenter_id is not None}
    machine_ids = {op.machine_id for op in operations if op.machine_id is not None}
    part_type_ids = {op.part_type_id for op in operations if op.part_type_id is not None}
    work_center_map = {}
    machine_map = {}
    part_type_map = {}
    if work_center_ids:
        work_centers = db.query(WorkCenterModel).filter(WorkCenterModel.id.in_(work_center_ids)).all()
        work_center_map = {wc.id: wc.work_center_name for wc in work_centers}
    if machine_ids:
        machines = db.query(MachineModel).filter(MachineModel.id.in_(machine_ids)).all()
        machine_map = {m.id: m.make for m in machines}
    if part_type_ids:
        part_types = db.query(PartTypeModel).filter(PartTypeModel.id.in_(part_type_ids)).all()
        part_type_map = {pt.id: pt.type_name for pt in part_types}

    for op in operations:
        op.work_center_name = work_center_map.get(op.workcenter_id)
        op.machine_name = machine_map.get(op.machine_id)
        op.part_type_name = part_type_map.get(op.part_type_id)

    return operations


@router.get("/{operation_id}", response_model=Operation)
def get_operation(operation_id: int, db: Session = Depends(get_db)):
    operation = (
        db.query(OperationModel)
        .options(joinedload(OperationModel.user))
        .filter(OperationModel.id == operation_id)
        .first()
    )
    if not operation:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Operation with id {operation_id} not found"
        )
    work_center = None
    if operation.workcenter_id is not None:
        work_center = db.query(WorkCenterModel).filter(WorkCenterModel.id == operation.workcenter_id).first()
    machine = None
    if operation.machine_id is not None:
        machine = db.query(MachineModel).filter(MachineModel.id == operation.machine_id).first()
    part_type = None
    if operation.part_type_id is not None:
        part_type = db.query(PartTypeModel).filter(PartTypeModel.id == operation.part_type_id).first()
    operation.work_center_name = work_center.work_center_name if work_center else None
    operation.machine_name = machine.make if machine else None
    operation.part_type_name = part_type.type_name if part_type else None
    return operation


@router.get("/part/{part_id}", response_model=List[Operation])
def get_operations_by_part(part_id: int, user_id: int | None = None, db: Session = Depends(get_db)):
    query = (
        db.query(OperationModel)
        .options(
            joinedload(OperationModel.user),
            joinedload(OperationModel.operation_documents),
            joinedload(OperationModel.tools).joinedload(ToolWithPartModel.tool)
        )
        .filter(OperationModel.part_id == part_id)
        .order_by(OperationModel.id.asc())
    )
    if user_id is not None:
        query = query.filter(OperationModel.user_id == user_id)
    operations = query.all()

    work_center_ids = {op.workcenter_id for op in operations if op.workcenter_id is not None}
    machine_ids = {op.machine_id for op in operations if op.machine_id is not None}
    part_type_ids = {op.part_type_id for op in operations if op.part_type_id is not None}
    work_center_map = {}
    machine_map = {}
    part_type_map = {}
    if work_center_ids:
        work_centers = db.query(WorkCenterModel).filter(WorkCenterModel.id.in_(work_center_ids)).all()
        work_center_map = {wc.id: wc.work_center_name for wc in work_centers}
    if machine_ids:
        machines = db.query(MachineModel).filter(MachineModel.id.in_(machine_ids)).all()
        machine_map = {m.id: m.make for m in machines}
    if part_type_ids:
        part_types = db.query(PartTypeModel).filter(PartTypeModel.id.in_(part_type_ids)).all()
        part_type_map = {pt.id: pt.type_name for pt in part_types}

    for op in operations:
        op.work_center_name = work_center_map.get(op.workcenter_id)
        op.machine_name = machine_map.get(op.machine_id)
        op.part_type_name = part_type_map.get(op.part_type_id)
        if op.tools:
            op.tools.sort(key=lambda x: x.id)

    return operations


@router.put("/{operation_id}", response_model=Operation)
def update_operation(operation_id: int, operation: OperationUpdate, db: Session = Depends(get_db)):
    db_operation = db.query(OperationModel).filter(OperationModel.id == operation_id).first()
    if not db_operation:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Operation with id {operation_id} not found"
        )

    update_data = operation.model_dump(exclude_unset=True)

    if "operation_number" in update_data:
        new_op_num_raw = update_data.get("operation_number")
        new_op_num = new_op_num_raw.strip() if isinstance(new_op_num_raw, str) else None
        if not new_op_num:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="operation_number cannot be empty",
            )
        existing = (
            db.query(OperationModel)
            .filter(
                OperationModel.part_id == db_operation.part_id,
                OperationModel.operation_number == new_op_num,
                OperationModel.id != operation_id,
            )
            .first()
        )
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Operation number '{new_op_num}' already exists for this part",
            )
        update_data["operation_number"] = new_op_num

    part_type_id = update_data.get("part_type_id") if "part_type_id" in update_data else db_operation.part_type_id
    if part_type_id == 2:
        from_date = update_data.get("from_date") if "from_date" in update_data else db_operation.from_date
        to_date = update_data.get("to_date") if "to_date" in update_data else db_operation.to_date
        if not from_date or not to_date:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Outsource operations require from_date and to_date",
            )

    zero_time = time(0, 0, 0)
    new_setup = update_data.get("setup_time") if "setup_time" in update_data else db_operation.setup_time
    new_cycle = update_data.get("cycle_time") if "cycle_time" in update_data else db_operation.cycle_time
    if part_type_id != 2:
        if (
            not new_setup
            or not new_cycle
            or new_setup == zero_time
            or new_cycle == zero_time
        ):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="setup_time and cycle_time are mandatory and cannot be 00:00:00 for non Out-Source operations",
            )
    for field, value in update_data.items():
        setattr(db_operation, field, value)

    db.commit()
    db.refresh(db_operation)
    db_operation = (
        db.query(OperationModel)
        .options(joinedload(OperationModel.user))
        .filter(OperationModel.id == operation_id)
        .first()
    )
    pt = db.query(PartTypeModel).filter(PartTypeModel.id == db_operation.part_type_id).first()
    db_operation.part_type_name = pt.type_name if pt else None
    return db_operation


@router.delete("/{operation_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_operation(operation_id: int, db: Session = Depends(get_db)):
    db_operation = db.query(OperationModel).filter(OperationModel.id == operation_id).first()
    if not db_operation:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Operation with id {operation_id} not found"
        )

    documents = db.query(OperationDocumentModel).filter(OperationDocumentModel.operation_id == operation_id).all()
    minio_client = get_minio_client()

    for doc in documents:
        try:
            if doc.document_url:
                parsed_url = urlparse(doc.document_url)
                path_parts = parsed_url.path.lstrip('/').split('/', 1)
                if len(path_parts) >= 2:
                    bucket_name = path_parts[0]
                    object_name = path_parts[1]
                    minio_client.client.remove_object(bucket_name, object_name)
                elif not parsed_url.netloc and '/' in doc.document_url:
                    path_parts = doc.document_url.lstrip('/').split('/', 1)
                    if len(path_parts) >= 2:
                        bucket_name = path_parts[0]
                        object_name = path_parts[1]
                        minio_client.client.remove_object(bucket_name, object_name)
        except Exception as e:
            print(f"Error deleting file from MinIO for document {doc.id}: {str(e)}")
        db.delete(doc)

    tools = db.query(ToolWithPartModel).filter(ToolWithPartModel.operation_id == operation_id).all()
    for tool in tools:
        db.delete(tool)

    db.delete(db_operation)
    db.commit()
    return None


@router.post("/parse-mpp", response_model=List[OperationPreview])
async def parse_mpp_file(file: UploadFile = File(...)):
    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Empty file",
        )
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext == ".csv":
        operations = _parse_csv(content)
    elif ext in (".xlsx", ".xls"):
        operations = _parse_xlsx(content)
    elif ext == ".docx":
        operations = _parse_docx(content)
    elif ext == ".pdf":
        operations = _parse_pdf(content)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Use DOCX, CSV, XLSX, or PDF.",
        )
    if not operations:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not extract operations from file. Make sure the file contains columns: Op Number, Operation Name, Setup Time, Cycle Time, Work Instructions, Notes.",
        )
    return operations